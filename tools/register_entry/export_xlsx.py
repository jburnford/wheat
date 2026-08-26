#!/usr/bin/env python3
"""Write rows entered in the app (data/register/*.csv, entered_at set) back into
W1/W2/W3.xlsx so the existing map/analysis pipeline sees them.

    python tools/register_entry/export_xlsx.py [--dry-run]

For each entered quarter the matching workbook row (same QSECT/PSECT/PTWP/PRGE)
gets Type, FirstDate, FirstDateSuccess, Successful_Claims, Failed_Claims,
Patent_Date, Number_of_claims, Notes replaced; spilled index-name cells in
Patent_Date / Number_of_claims / Notes / Unnamed columns are cleared on those
rows. A timestamped backup of each workbook is written first.
"""
import argparse, csv, shutil
from pathlib import Path
from datetime import datetime
import openpyxl

ROOT = Path(__file__).resolve().parents[2]; DATA = ROOT / "data" / "register"
REG = ["Type", "FirstDate", "FirstDateSuccess", "Successful_Claims", "Failed_Claims", "Patent_Date", "Number_of_claims", "Notes"]

def main():
    ap = argparse.ArgumentParser(); ap.add_argument("--dry-run", action="store_true"); a = ap.parse_args()
    by_book = {}
    for p in sorted(DATA.glob("W*_R*.csv")):
        mer, sheet = p.stem.split("_")
        with open(p, newline="", encoding="utf-8") as f:
            rows = [r for r in csv.DictReader(f) if r.get("entered_at")]
        if rows: by_book.setdefault(mer, {})[sheet] = rows
    if not by_book: print("nothing entered in the app yet"); return
    for mer, sheets in by_book.items():
        book = ROOT / f"{mer}.xlsx"
        if not a.dry_run:
            bak = ROOT / f"{mer}.backup-{datetime.now():%Y%m%d-%H%M}.xlsx"; shutil.copy(book, bak); print("backup:", bak.name)
        wb = openpyxl.load_workbook(book)
        total = 0
        for sheet, rows in sheets.items():
            ws = wb[sheet]
            hdr = {str(c.value).strip(): i + 1 for i, c in enumerate(ws[1]) if c.value is not None}
            need = ["QSECT", "PSECT", "PTWP", "PRGE"] + REG
            missing = [c for c in need if c not in hdr]
            if missing: print(f"{mer} {sheet}: missing columns {missing}, skipped"); continue
            spill_cols = [i for h, i in hdr.items() if h.startswith("Unnamed")]
            index = {}
            for r in range(2, ws.max_row + 1):
                try: k = (str(ws.cell(r, hdr["QSECT"]).value).strip().upper(), int(ws.cell(r, hdr["PSECT"]).value), int(ws.cell(r, hdr["PTWP"]).value), int(ws.cell(r, hdr["PRGE"]).value))
                except (TypeError, ValueError): continue
                index[k] = r
            n = 0
            for row in rows:
                k = (row["QSECT"].upper(), int(row["PSECT"]), int(row["PTWP"]), int(row["PRGE"])); r = index.get(k)
                if r is None: print(f"  {mer} {sheet}: no workbook row for {k}"); continue
                for c in REG:
                    v = row.get(c, "")
                    if c == "Number_of_claims" and v.isdigit(): v = int(v)
                    ws.cell(r, hdr[c]).value = v if v != "" else None
                for i in spill_cols: ws.cell(r, i).value = None
                n += 1
            print(f"{mer} {sheet}: {n} rows written"); total += n
        if not a.dry_run:
            wb.save(book); print(f"saved {book.name} ({total} rows)")
        else:
            print(f"dry run: would write {total} rows to {book.name}")

if __name__ == "__main__":
    main()
